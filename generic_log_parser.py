"""
Author:Michael Z
Date: 14-Mar-2023
Python Version: 3.9.10

Generic Log Parser is used to parse one or more log files and export to a single .xlsx file.
For multiple log files, each log file is exported to a separate sheet in the same xlsx document.

Most log files should be parsed using parse_all() and customising the DEFAULT_DELIMS
constant.

If your log files are formatted horribly use the parse_labels and parse_data functions.
These functions allow for more precise separation of labels and data (TBC - 26/04/2023).

TODO: First version of this application will be run directly as a single script.
      Future versions could have GUI and be compiled into executable for ease of use.
"""

import multiprocessing as mp
import logging as log
import tkinter as tk
from tkinter import filedialog as fd
import openpyxl as xlsx
import time
import os

# Number of CPU cores to reserve for other tasks external to this application.
RESERVE_CORES = 1

# If parsing more than EFFICIENCY_THRESH number of log files, process using multiprocessing.
# Otherwise if parsing less log files, it is more efficient to parse sequencially.
EFFICIENCY_THRESH = 5

# Section keys used to denote the last line of text in a log file "section" where formatting
# is very different in the following section. These should be unique words or phrases. Tuple
# items are in order of the section number. i.e. SECTION_KEYS[0] relates to the end of
# the first section in the log file. The last section at the end of the log file does not need a
# section delimiter.
SECTION_KEYS = (" 2023", "Device Chemistry")
SECTION_DELIMS = [(" ", ","), ("=",)]

# Delimiters used to separate data and labels. Used for the parse_all function.
# DEFAULT_DELIMS = (",", ":", " ", "=")
DEFAULT_DELIMS = (",")

# Delimiters used to isolate labels. Used for the parse_labels function.
LABEL_DELIMS = ("-", ":")

# Delimiters used to isolate labels. Used for the parse_labels function.
DATA_DELIMS = (" ", ",")

# Output .xlsx file name.
OUTPUT_FILE_NAME = "Parsed_Logs.xlsx"


def parse(line: str, delims: tuple) -> list:
    """
    Parse an individual line from a log file.

    Args:
        line: Individual line of a log file.
        delims: Delimiters used to separate labels and data.

    Returns:
        List of parsed labels and data with leading/trailing whitespace removed.

    """
    # Replace and split is faster than regex split method.
    for delim in delims:
        if delim == delims[0]:
            pass
        else:
            line = line.replace(delim, delims[0])
    ret = line.split(delims[0])
    return [x.strip() for x in ret]


def parse_all(log_file=None, delims=None, section_keys=None):
    """
    Parse everything in a log file. return a list of lists where each "sub-list"
    represents a parsed line in the log file.

    Args:
        log_file: Full path of the log file to be parsed.
        delims: Tuple conaining delimiter characters. If multiple sections, this will be a list
            of tuples.
        section_keys: Tuple containing unique words or phrases that denote the last line of a
            section within the log file.

    Returns:
        List of lists where each "sub-list" represents a parsed line in the log file.

    """
    ret = []
    with open(log_file, 'r') as lf:
        file = lf.readlines()
        if section_keys is None:
            # No separate sections.
            for line in file:
                ret.append(parse(line, delims))
        else:
            # Parse each section of the log file separately.
            section_num = 0
            for line in file:
                ret.append(parse(line, delims[section_num]))
                if section_num < len(section_keys):
                    if section_keys[section_num] in line:
                        section_num += 1
    return ret


def export_xlsx(log_files, parsed_lists):
    """
    Export parsed content into an xlsx spreadsheet. Multiple parsed files can be exported
    onto separate sheets of the xlsx file.
    """
    # TODO: Much of this function is to be updated to be more user friendly
    # such as more flexible output path, different output fields, etc.
    # FIXME: fix issue where all cells are formatted as text in the final xlsx doc.
    # Requires formatting as numbers to allow user to start plotting/analysing data in excel.
    log_path = os.path.dirname(log_files[0])
    save_file = os.path.join(log_path, OUTPUT_FILE_NAME)
    existing_xlsx = False

    # check if existing workbook and add to that.
    # Assuming file would be in same folder as this script.
    if os.path.exists(save_file):
        existing_xlsx = True
        wb = xlsx.load_workbook(filename=save_file)
    else:
        wb = xlsx.Workbook()
        wb.active.title = "Sheet 1"

    sheet_names = wb.sheetnames
    sheet_num = len(sheet_names)
    sheet = wb[sheet_names[-1]]

    for idx, file in enumerate(log_files):
        log_name = os.path.basename(file)
        # Create new sheet if editing pre-existing workbook.
        if idx > 0 or existing_xlsx:
            sheet_num += 1
            sheet = wb.create_sheet(f"Sheet {sheet_num}")

        sheet["A1"] = (log_name)
        sheet["A2"] = "--------"

        for line in parsed_lists[idx]:
            sheet.append(line)

    wb.save(save_file)
    return


def parse_labels(delim=None):
    """
    Parse only data labels.
    """
    if delim is None:
        delim = DEFAULT_DELIMS

    # TODO: Finish at some later date.


def parse_data(delim=None):
    """
    Parse only data values.
    """
    if delim is None:
        delim = DEFAULT_DELIMS

    # TODO: Finish at some later date.


if __name__ == '__main__':
    log.basicConfig(level=log.DEBUG)

    root = tk.Tk()
    root.withdraw()
    folder = fd.askdirectory(title="Choose a folder containing log files")

    if not folder:
        log.info("No folder selected.")
        time.sleep(2)
        quit()

    files = os.walk(folder)
    log_files = []
    results = []

    for root, directories, filenames in os.walk(folder):
        for filename in filenames:
            if filename.endswith(".log") or filename.endswith(".txt"):
                log.debug(filename)
                log_files.append(os.path.join(root, filename))

        log.info(f"Number of log files to parse: {len(log_files)}")

    # Based on number of files to be parsed, determine the most efficient processing method.
    if len(log_files) > EFFICIENCY_THRESH:
        args = []
        cpu_cores = mp.cpu_count()
        log.debug(f"Total no. of CPU cores / logical processors): {cpu_cores}")
        start_time = time.time()
        with mp.Pool(cpu_cores - RESERVE_CORES) as pool:
            for file in log_files:
                args.append((file, DEFAULT_DELIMS))
            for result in pool.starmap(parse_all, args):
                results.append(result)
    else:
        start_time = time.time()
        for file in log_files:
            results.append(parse_all(file, DEFAULT_DELIMS))

    total_time = time.time() - start_time
    log.debug(f"Total Parsing Time = {total_time:.6f}")

    export_xlsx(log_files, results)
